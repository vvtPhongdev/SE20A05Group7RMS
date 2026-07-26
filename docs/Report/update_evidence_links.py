import openpyxl
import re

def parse_links_file(file_path):
    links = {}
    current_week = None
    
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
                
            week_match = re.match(r"^\[(Week\d+)\]", line)
            if week_match:
                current_week = week_match.group(1)
                links[current_week] = {}
                continue
                
            task_match = re.match(r"^Task (\d+):\s*(.*)", line)
            if task_match and current_week:
                task_num = int(task_match.group(1))
                link_url = task_match.group(2).strip()
                links[current_week][task_num] = link_url
                
    return links

def update_excel_links(excel_path, links_dict):
    wb = openpyxl.load_workbook(excel_path)
    
    updated_count = 0
    for week_name, tasks in links_dict.items():
        if week_name not in wb.sheetnames:
            print(f"  Warning: sheet {week_name} not found in workbook.")
            continue
            
        sheet = wb[week_name]
        
        for task_num, link_url in tasks.items():
            r = task_num + 1 # row index is task_num + 1 (since row 2 is task 1)
            
            if r <= sheet.max_row:
                cell = sheet.cell(r, 7) # column 7 is Evidence / Link
                
                # Check if it was different
                old_val = cell.value
                if old_val != link_url:
                    cell.value = link_url
                    # We also assign the hyperlink property so it remains clickable in Excel
                    cell.hyperlink = link_url
                    print(f"  [{week_name}] Row {r} (Task {task_num}): Link updated from '{old_val}' to '{link_url}'")
                    updated_count += 1
                    
    if updated_count > 0:
        wb.save(excel_path)
        print(f"Excel updated successfully. Total links modified: {updated_count}")
    else:
        print("No link changes found.")

if __name__ == "__main__":
    links_file = "docs/Report/links_to_update.txt"
    excel_file = "docs/Report/SWP_SE20A05_Group7_ReportAI.xlsx"
    
    try:
        links_dict = parse_links_file(links_file)
        update_excel_links(excel_file, links_dict)
    except Exception as e:
        print(f"Error updating links: {e}")
