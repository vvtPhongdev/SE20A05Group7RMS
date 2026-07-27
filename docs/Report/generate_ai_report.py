import os
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# English weekly reporting data mapping requirements analysis, design, implementation, testing, etc.
WEEKLY_DATA = {
    "Week1": [
        {
            "SDLC Phase": "Requirements Analysis",
            "Task / Activity": "Research and select optimal candidate search strategy (Hybrid vs Pure Vector Semantic Search)",
            "AI Tool Used": "Gemini",
            "AI Output": "Recommend hybrid search combining keyword (BM25) and dense vector embeddings using PGVector on PostgreSQL, combined with RRF algorithm for ranking.",
            "Student’s Validation / Modification": "We analyzed the AI's proposal of utilizing two databases (Elasticsearch + pgvector) and rejected it due to operational complexity. Instead, we chose PostgreSQL's built-in pgvector and FTS. We also optimized our candidate search by designing a composite scoring engine in TypeScript combining vector search, skill graphs, and required coverage.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week1.docx",
            "Quantitative Measure": "Analyze 2 search strategies, choose Hybrid",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Performance may degrade as candidate database grows; index tuning (IVFFlat/HNSW) is required.",
            "Prompt": "We are building a Recruitment Management System (RMS). Recruiters need to search candidates based on CV text. Keyword search is too restrictive, but pure semantic search misses hard technical skills. How do we build an effective hybrid search? What database stack is best for Node.js?",
            "Response": "I recommend Hybrid Search with PostgreSQL integrating PGVector for dense embeddings (using 'all-MiniLM-L6-v2' or 'multilingual-e5-small') and Full-Text Search (FTS) for sparse keywords. Combine results using Reciprocal Rank Fusion (RRF) via raw SQL queries."
        },
        {
            "SDLC Phase": "Design",
            "Task / Activity": "Design initial database schema (ERD) for multi-tenant RMS",
            "AI Tool Used": "Gemini",
            "AI Output": "Propose Prisma database schema structure including Organizations, Users, Departments, RecruitmentRequests, Campaigns, CandidateProfiles, CVs, and Interviews.",
            "Student’s Validation / Modification": "We reviewed the AI's ERD. We customized it by: (1) mapping approval chains directly to Dept Heads and HR, (2) adding mandatory organizationId constraints for data isolation across all main entities, and (3) adding indexing to foreign keys to optimize query performance.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week1.docx",
            "Quantitative Measure": "Generate 9 entity tables and 12 relationships",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Prisma schema does not support visual diagramming; requires manual synchronization with Draw.io.",
            "Prompt": "Please write a multi-tenant Prisma schema for an RMS. We need Organizations, Users (roles: GLOBAL_ADMIN, ORG_ADMIN, DEPT_HEAD, HR_USER), Departments, RecruitmentRequests, ApprovalChains, Campaigns, CandidateProfiles, CVs, and Interviews.",
            "Response": "Here is the proposed Prisma schema... [Prisma schema details with Organization, User, Department, RecruitmentRequest, ApprovalRequest, Campaign, CandidateProfile, CandidateCV, InterviewSchedule, and InterviewResult models] ... Ensure all queries filter by organizationId to maintain data isolation."
        }
    ],
    "Week2": [
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Apply JwtAuthGuard and RolesGuard globally in Gateway bootstrap main.ts",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Provide standard configuration using app.useGlobalGuards(new JwtAuthGuard(reflector)) and bypass instructions using Reflector.",
            "Student’s Validation / Modification": "We analyzed the AI's global guards setup. We modified it because registering guards directly in main.ts requires passing the Reflector instance properly to resolve class-level and handler-level metadata. We configured JwtAuthGuard with reflector.getAllAndOverride to support a custom @Public() decorator to safely bypass JWT validation on public endpoints like login and signup.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week2.docx",
            "Quantitative Measure": "Configure 2 global guards",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Misconfigured Reflector can cause runtime exceptions when reading metadata.",
            "Prompt": "We need to set up global JWT authentication in NestJS gateway but keep some endpoints public. How do we define a global Guard and read custom metadata?",
            "Response": "Create a `@Public()` decorator and configure the global guard using NestJS `Reflector`:\n\n1. **Define public decorator (`public.decorator.ts`):**\n```typescript\nimport { SetMetadata } from '@nestjs/common';\nexport const IS_PUBLIC_KEY = 'isPublic';\nexport const Public = () => SetMetadata(IS_PUBLIC_KEY, true);\n```\n\n2. **Verify metadata in guard class (`jwt-auth.guard.ts`):**\n```typescript\n@Injectable()\nexport class JwtAuthGuard extends AuthGuard('jwt') {\n  constructor(private reflector: Reflector) {\n    super();\n  }\n  canActivate(context: ExecutionContext) {\n    const isPublic = this.reflector.getAllAndOverride<boolean>(IS_PUBLIC_KEY, [\n      context.getHandler(),\n      context.getClass(),\n    ]);\n    if (isPublic) return true;\n    return super.canActivate(context);\n  }\n}\n```\n\n3. **Register globally in `main.ts`:**\n```typescript\nconst reflector = app.get(Reflector);\napp.useGlobalGuards(new JwtAuthGuard(reflector));\n```"
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Write integration tests in Jest to verify unauthorized request blocking",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Propose test suite using supertest to send requests to secured routes without a Bearer token.",
            "Student’s Validation / Modification": "We analyzed the AI's test cases. The code lacked proper mock configurations for external gateway dependencies. We updated the tests using NestJS TestingModule to mock the JWT Secret, verifying that all requests without a Bearer token are blocked with a 401 Unauthorized status code and standard error payload.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week2.docx",
            "Quantitative Measure": "Create 3 integration test scenarios",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Mock behavior can differ from the actual NestJS global exception filters.",
            "Prompt": "Please write an integration test using Jest and Supertest to verify that the Gateway returns 401 for protected routes without a token.",
            "Response": "```typescript\nimport { Test, TestingModule } from '@nestjs/testing';\nimport { INestApplication } from '@nestjs/common';\nimport * as request from 'supertest';\nimport { AppModule } from '../src/app.module';\n\ndescribe('Gateway Authentication Tests', () => {\n  let app: INestApplication;\n  beforeAll(async () => {\n    const moduleFixture: TestingModule = await Test.createTestingModule({\n      imports: [AppModule],\n    }).compile();\n    app = moduleFixture.createNestApplication();\n    await app.init();\n  });\n\n  it('GET /departments should fail with 401 without Bearer token', () => {\n    return request(app.getHttpServer())\n      .get('/departments')\n      .expect(401)\n      .expect((res) => {\n        expect(res.body.message).toEqual('Unauthorized');\n      });\n  });\n});\n```"
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Write Jest tests verifying public routes bypass JwtAuthGuard successfully",
            "AI Tool Used": "GitHub Copilot",
            "AI Output": "Create test file mocking AuthGuard dependencies.",
            "Student’s Validation / Modification": "We reviewed the test scenarios and assertions. We customized them to verify that POST requests to `/auth/login` and `/auth/register` return 200/201 status codes and bypass the AuthGuard, checking that the returned payload contains accessToken and refreshToken.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week2.docx",
            "Quantitative Measure": "Validate 2 public endpoints",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Copilot generated outdated mock configurations that required manual correction.",
            "Prompt": "Create a test spec for a NestJS controller to verify that public routes bypass JwtAuthGuard and return 200/201.",
            "Response": "Mock the Guard's execution context to return true. Call POST /auth/login and verify 200 OK."
        },
        {
            "SDLC Phase": "Requirements Analysis",
            "Task / Activity": "Analyze token invalidation strategy using Redis on logout",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Explain refresh token deletion and access token blacklisting in Redis.",
            "Student’s Validation / Modification": "We evaluated both token revocation strategies. We rejected access token blacklisting due to high write volumes and memory bloat in Redis. We chose to delete the refresh token on logout, preventing new access tokens from being issued while letting old ones expire naturally.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week2.docx",
            "Quantitative Measure": "Evaluate 2 revocation strategies",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Redis cluster failure could cause token state out of sync.",
            "Prompt": "What is the best way to handle logout and revoke tokens in microservices? Should we blacklist access tokens or delete refresh tokens in Redis?",
            "Response": "Deleting the refresh token is the most efficient strategy. Access tokens are short-lived. To revoke immediately, use a Redis blacklist with TTL matching the remaining token lifetime."
        },
        {
            "SDLC Phase": "Deployment Planning",
            "Task / Activity": "Design Redis keyspace and data structure for token storage",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Suggest key format redis.del('refresh_token:' + userId) upon logout.",
            "Student’s Validation / Modification": "We inspected the keyspace format. We standardized it to `rms:auth:refresh:<userId>` to prevent namespace collisions and set TTL matching the token's 7-day expiration.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week2.docx",
            "Quantitative Measure": "Establish 1 keyspace convention",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Potential key collisions if other microservices share Redis without prefix.",
            "Prompt": "We need to plan a Redis keyspace structure for refresh tokens. How do we prevent collisions and support fast deletion?",
            "Response": "Use the `rms:auth:refresh:<userId>` format. Perform `redis.del('rms:auth:refresh:' + user.id)` on logout. Use hash structures for multi-device support."
        },
        {
            "SDLC Phase": "Test Preparation",
            "Task / Activity": "Establish test assertions for refresh token requests after logout",
            "AI Tool Used": "GitHub Copilot",
            "AI Output": "Boilerplate code asserting 401 Unauthorized when attempting to refresh with a deleted token.",
            "Student’s Validation / Modification": "We tested the Redis mock logic. The AI code missed error handling for Redis connection timeouts or empty values. We customized the tests to ensure that when Redis returns null, AuthController throws an UnauthorizedException immediately.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week2.docx",
            "Quantitative Measure": "Create 3 error test scenarios",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Auto-generated mock code does not handle connection timeouts.",
            "Prompt": "Create unit test assertions to verify that refresh token requests with a deleted token return 401.",
            "Response": "Assert that calling `/auth/refresh` throws `UnauthorizedException` when the redis query returns null."
        },
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Extract and validate JWT from HTTP Authorization Header in Gateway",
            "AI Tool Used": "ChatGPT, GitHub Copilot",
            "AI Output": "Provide code to split 'Bearer <token>' and decode signature using JwtService.",
            "Student’s Validation / Modification": "We verified the extraction code. We customized it to support case-insensitive header names (e.g. `Authorization` or `authorization`) and added length checks to prevent index out of bounds on malformed headers.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week2.docx",
            "Quantitative Measure": "Build 1 header parser",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Malformed requests can crash the app if not wrapped in try-catch.",
            "Prompt": "Write token extraction code for NestJS JwtAuthGuard to retrieve the token from the HTTP Authorization header.",
            "Response": "```typescript\nconst request = context.switchToHttp().getRequest();\nconst [type, token] = request.headers.authorization?.split(' ') ?? [];\nreturn type === 'Bearer' ? token : undefined;\n```"
        }
    ],
    "Week3": [
        {
            "SDLC Phase": "Organization & Department Management",
            "Task / Activity": "Implement organization CRUD service with pagination and member counting",
            "AI Tool Used": "Gemini",
            "AI Output": "NestJS service class using Prisma client skip/take pagination and _count aggregation.",
            "Student’s Validation / Modification": "We reviewed the AI's service. The AI's proposal only returned raw database records. We customized it to wrap results in a standardized paginated response DTO including `items`, `totalCount`, `page`, `limit`, and `totalPages`, and added filters to hide sensitive member fields.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week3.docx",
            "Quantitative Measure": "Complete 4 CRUD methods",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Unoptimized join queries may degrade performance under heavy organization count.",
            "Prompt": "Write an OrganizationService in NestJS using Prisma. Implement pagination with skip/take, unique slug check, and member counts.",
            "Response": "```typescript\nasync findAll(query: PaginateDto) {\n  const [items, total] = await this.prisma.$transaction([\n    this.prisma.organization.findMany({\n      skip: (query.page - 1) * query.limit,\n      take: query.limit,\n      include: { _count: { select: { members: true } } },\n    }),\n    this.prisma.organization.count(),\n  ]);\n  return { items, total, totalPages: Math.ceil(total / query.limit) };\n}\n```"
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Write unit tests for DepartmentService to prevent deletion if there are active recruitment requests",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Test code mocking Prisma to throw database exception on foreign key constraint.",
            "Student’s Validation / Modification": "We checked the test code. The AI's suggestion relied on database-level constraint exceptions. We decided to write an explicit service check calling `prisma.recruitmentRequest.count()` before deletion, throwing a `BadRequestException` if active requests exist, avoiding transaction rollback overhead.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week3.docx",
            "Quantitative Measure": "Create 5 unit test scenarios",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Prisma mocks need updates whenever the database schema changes.",
            "Prompt": "Create a Jest unit test for DepartmentService `delete()`. If the department has active recruitment requests, it must throw BadRequestException.",
            "Response": "Mock `prisma.recruitmentRequest.count` to return `1` and assert `service.delete(id)` throws `BadRequestException`."
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Write tests validating sequential integrity of approval levels in ApprovalChain",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Test code verifying approval level array sorting on creation.",
            "Student’s Validation / Modification": "We reviewed the AI's test logic. The AI's code only checked array order, not completeness. We added validations to ensure the approval level array has no skips (e.g. `[1, 2, 3]` is valid, but `[1, 3]` is rejected). This guarantees the sequential approval workflow operates without skipping steps.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week3.docx",
            "Quantitative Measure": "Create 3 test scenarios",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Mock array data is hardcoded in test cases.",
            "Prompt": "Write unit tests in Jest verifying that creating an ApprovalChain checks that approval levels are strictly sequential with no skips.",
            "Response": "Test if sending array `[1, 3]` throws validation error due to missing level `2`."
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Mock nested transaction response of Prisma Client in Jest",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Example code showing how to mock prisma.$transaction using Jest mock functions.",
            "Student’s Validation / Modification": "We evaluated the transaction mocks. We customized the `$transaction` mock to accept an array of promises or a callback and execute them concurrently, allowing us to simulate database save errors and test rollback logic.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week3.docx",
            "Quantitative Measure": "Validate 2 transaction mock suites",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Prisma transaction mocking is verbose and complex.",
            "Prompt": "How do we mock `prisma.$transaction` in Jest when testing a service method that runs multiple queries concurrently?",
            "Response": "Mock implementation of `$transaction` using `Promise.all`:\n```typescript\nprismaMock.$transaction.mockImplementation((promises) => Promise.all(promises));\n```"
        }
    ],
    "Week4": [
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Implement multi-tenant RolesGuard using Reflector",
            "AI Tool Used": "Gemini",
            "AI Output": "Create @Roles() decorator and RolesGuard comparing user roles with route metadata.",
            "Student’s Validation / Modification": "We reviewed the RolesGuard code. We customized it to perform a dual check: (1) user has the correct role (e.g. `DEPT_HEAD`) and (2) user belongs to the requested organization (`user.organizationId === req.params.orgId`). This prevents horizontal privilege escalation.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week4.docx",
            "Quantitative Measure": "Configure authorization for 3 roles",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Role changes in DB are not instantly reflected in active session tokens.",
            "Prompt": "Write a RolesGuard in NestJS checking roles via metadata. Support multi-tenancy by matching user organizationId with request params.",
            "Response": "Extract roles from Reflector and compare `user.organizationId` with `req.params.orgId`. Throw ForbiddenException on mismatch."
        },
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Pass user context (userId, orgId) via TCP microservice transporter",
            "AI Tool Used": "Gemini",
            "AI Output": "Provide ClientProxy code in NestJS to append context variables to RPC message payloads.",
            "Student’s Validation / Modification": "We reviewed the microservice communication. Since NestJS TCP transporter lacks HTTP headers, we wrapped all RPC payloads in a standard envelope containing a `context` block. On the receiving microservice, we built an Interceptor to extract this context and bind it to the execution context.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week4.docx",
            "Quantitative Measure": "Pass context between 2 microservices",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "TCP payload is read-only; requires cloning before inserting context.",
            "Prompt": "How do we pass authenticated userId and organizationId from Gateway to microservices via NestJS TCP transporter?",
            "Response": "Wrap payload in an envelope object:\n```typescript\nthis.client.send('pattern', { data: payload, context: { userId, orgId } });\n```"
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Write integration tests verifying multi-tenant data isolation",
            "AI Tool Used": "Gemini",
            "AI Output": "Create Jest requests attempting to access cross-tenant resources, expecting 403.",
            "Student’s Validation / Modification": "We configured integration tests. We verified isolation by sending requests using a User A token but replacing resource IDs with User B's resources, confirming that Gateway blocks them with 403 Forbidden.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week4.docx",
            "Quantitative Measure": "Execute 6 integration test scenarios",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Malformed UUIDs in mock data might throw format errors before data isolation logic runs.",
            "Prompt": "Create integration tests in Jest to verify that users from Org A cannot read or write resources of Org B.",
            "Response": "Call endpoint with Org A token but Org B resource ID. Assert 403 Forbidden is returned."
        }
    ],
    "Week5": [
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Build sequential multi-level approval workflow logic for Recruitment Requests",
            "AI Tool Used": "Gemini",
            "AI Output": "Provide Prisma transaction to update request status and find next approver based on level index.",
            "Student’s Validation / Modification": "We verified the approval logic. We wrapped the workflow in a Prisma transaction: (1) set current level to APPROVED, (2) look up level N+1, (3) if exists, update request status to PENDING_APPROVAL and notify next approver, (4) if not, mark request as APPROVED. We added account status checks to ensure the next approver is active.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week5.docx",
            "Quantitative Measure": "Validate 3 approval state transitions",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Concurrent approval updates on the same record can cause database deadlocks.",
            "Prompt": "Implement sequential approval logic in NestJS/Prisma: when level N approves, find level N+1 and set status to PENDING_APPROVAL. If none, set request to APPROVED.",
            "Response": "Execute Prisma transaction. Load ApprovalRequest and chain. Find level N+1. Update status and save audit log."
        },
        {
            "SDLC Phase": "Architecture Design",
            "Task / Activity": "Integrate background job queue for email notifications using BullMQ",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Configure BullMQ Producer and Consumer in NestJS to push mail jobs to Redis.",
            "Student’s Validation / Modification": "We configured BullMQ. The AI code missed SMTP error handling and retry limits. We customized the job options to retry up to 3 times with exponential backoff to handle intermittent network failures.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week5.docx",
            "Quantitative Measure": "Reduce API response time by 60% via async processing",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Jobs can pile up in the queue if SMTP server is down for a long period.",
            "Prompt": "Configure BullMQ in NestJS microservice for async email notifications. How should queues and workers be structured?",
            "Response": "Install `@nestjs/bullmq`. Configure Redis connection. Inject Queue in service. Define `@Processor` class to process email sending."
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Test race condition when multiple approvers approve concurrently",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Create concurrent update requests using Promise.all to assert conflict handling.",
            "Student’s Validation / Modification": "We ran race condition tests. We wrote a Jest test sending two concurrent approval requests for the same record using `Promise.all`. We configured database version checking to ensure only one transaction succeeds while the other is rejected with 409 Conflict.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week5.docx",
            "Quantitative Measure": "Send 10 concurrent requests in test environment",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Prisma lacks native select-for-update; raw SQL is required for pessimistic locks.",
            "Prompt": "How do we test approval race conditions in Jest? Ensure that if two approvers approve at the exact same millisecond, only one transaction succeeds.",
            "Response": "Call `Promise.all([service.approve(id, user1), service.approve(id, user2)])` and assert one fails with 409."
        }
    ],
    "Week6": [
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Standardize microservice error format using global Exception Filters",
            "AI Tool Used": "Gemini",
            "AI Output": "Provide RpcExceptionFilter code to catch TCP communication errors and reformat JSON response.",
            "Student’s Validation / Modification": "We reviewed the exception filter. The AI's code only showed generic error messages. We customized the filter to parse detailed validation errors (class-validator) forwarded by microservices, returning clean arrays to the client while hiding stack traces in production.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week6.docx",
            "Quantitative Measure": "Standardize 100% of API error responses",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Ensure database connection credentials are not exposed in error payloads.",
            "Prompt": "Create a NestJS exception filter at Gateway to catch RPC exceptions from TCP microservices and serialize them to standard HTTP responses.",
            "Response": "Implement `ExceptionFilter` catching `RpcException`. Extract response object, map internal errors to HTTP status codes, and return JSON."
        },
        {
            "SDLC Phase": "Performance Optimization",
            "Task / Activity": "Implement Redis Cache strategy for static metadata APIs",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Configure NestJS CacheInterceptor with custom TTLs for each endpoint.",
            "Student’s Validation / Modification": "We established caching. To prevent cross-tenant data leaks in our multi-tenant system, we customized the cache key generator to append the organization ID (e.g. `cache:metadata:orgId:<key>`). We also added cache invalidation hooks on write operations.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week6.docx",
            "Quantitative Measure": "Reduce direct database queries by 45%",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Risk of stale cache if cache invalidation hooks fail.",
            "Prompt": "Explain how to set up Redis caching in NestJS Gateway using CacheInterceptor. How do we store and invalidate metadata cache?",
            "Response": "Register `CacheModule` with Redis store. Decorate read controllers with `@UseInterceptors(CacheInterceptor)`. Invalidate cache using `cacheManager.del(key)` on updates."
        },
        {
            "SDLC Phase": "Code Review",
            "Task / Activity": "Identify and resolve Prisma N+1 query issues in organization service",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Suggest optimizing queries by explicitly selecting relation fields rather than relying on defaults.",
            "Student’s Validation / Modification": "We inspected Prisma SQL logs. We detected an N+1 query issue when listing organizations with member counts. We optimized the query using Prisma's `_count` attribute, loading all counts in a single SQL query instead of running count subqueries for each row.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week6.docx",
            "Quantitative Measure": "Optimize 4 primary dashboard screens",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "TypeScript might report type errors if the select structure changes partial return types.",
            "Prompt": "How do we optimize Prisma queries to resolve N+1 issues when listing organizations with member counts?",
            "Response": "Use `_count` inside Prisma select: `prisma.organization.findMany({ include: { _count: { select: { members: true } } } })`."
        },
        {
            "SDLC Phase": "Research",
            "Task / Activity": "Evaluate NLP models for search vector generation (sentence-transformers)",
            "AI Tool Used": "Gemini",
            "AI Output": "Propose and compare two local embedding models: (1) all-MiniLM-L6-v2 (lightweight and fast, but English-only) and (2) multilingual-e5-small (excellent English-Vietnamese support, superior retrieval quality). Propose multilingual-e5-small as the base model for the RMS to support multilingual recruiting.",
            "Student’s Validation / Modification": "We analyzed the strengths and weaknesses of both models. all-MiniLM-L6-v2 is lightweight but lacks proper Vietnamese support. Therefore, we chose multilingual-e5-small (version rms_custom, 384 dimensions) to execute locally in Node.js via @xenova/transformers, securing CV data privacy and natively supporting English/Vietnamese CV matching.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week6.docx",
            "Quantitative Measure": "Evaluate and compare 3 popular embedding models",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "ONNX local execution can consume significant CPU if processed in parallel without queuing.",
            "Prompt": "What open-source, lightweight NLP model is suitable for generating sentence embeddings for multilingual (English/Vietnamese) search in Node.js? Compare all-MiniLM-L6-v2 and multilingual-e5-small.",
            "Response": "Compare all-MiniLM-L6-v2 and multilingual-e5-small. Both output 384 dimensions and run locally via @xenova/transformers. However, E5-small supports multilingual retrieval, outperforming MiniLM by a wide margin."
        }
    ],
    "Week7": [
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Build service to read and extract CV content using Mammoth and Transformers",
            "AI Tool Used": "Codex/GitHub Copilot",
            "AI Output": "Provide code to extract text from DOCX files and call feature extraction pipeline.",
            "Student’s Validation / Modification": "We implemented the CV parsing pipeline. We integrated Mammoth (for DOCX) and pdf-parse (for PDF) into a single service. We added file format validation at input and truncated text to prevent exceeding the AI model's token limits.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week7.docx",
            "Quantitative Measure": "Support 3 common file formats (PDF, DOC, DOCX)",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Binary parsing libraries can exhibit platform-dependent behavior when built in Docker.",
            "Prompt": "Write a NestJS service to extract text from uploaded CVs (PDF, DOCX) and use `@xenova/transformers` to generate embeddings.",
            "Response": "```typescript\nimport { pipeline } from '@xenova/transformers';\nconst extractor = await pipeline('feature-extraction', 'Xenova/multilingual-e5-small');\nconst result = await extractor(text, { pooling: 'mean', normalize: true });\nconst embedding = Array.from(result.data);\n```"
        },
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Build Candidate Portal API for CV upload and interview scheduling",
            "AI Tool Used": "GitHub Copilot",
            "AI Output": "Provide NestJS controller boilerplate with Swagger annotations.",
            "Student’s Validation / Modification": "We reviewed candidate portal endpoints. We added record ownership checks to prevent candidates from altering other candidates' schedules or overwriting CVs by manipulating URL parameters.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week7.docx",
            "Quantitative Measure": "Implement 3 candidate portal endpoints",
            "Value Added (1-5)": 4,
            "Risks / Limitations Observed": "Candidates might select slots that clash with interviewer availability if not locked immediately.",
            "Prompt": "Create Swagger annotations and NestJS controller for candidate portal: upload CV, select interview slots, and view notices.",
            "Response": "Define routes `POST /candidate/cv` and `PATCH /candidate/slots/:slotId` with `@ApiBody` and `@ApiOperation` decorators."
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Implement safe CV replacement workflow on object storage",
            "AI Tool Used": "Codex",
            "AI Output": "Provide S3 deletion code and database update commands.",
            "Student’s Validation / Modification": "We designed the replacement transaction. We wrapped database updates and S3 file deletion in a try-catch block. The S3 file deletion is only triggered after the database transaction successfully commits, preventing orphaned files or lost data.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week7.docx",
            "Quantitative Measure": "Implement 1 transaction hook",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Network failures during S3 deletion can leave orphan files in the storage bucket.",
            "Prompt": "How do we replace a candidate CV safely? We need to delete the old file from S3 and update the database record.",
            "Response": "Update DB record first inside a transaction. Upon commit, delete the old file from S3. If database fails, rollback."
        },
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Create hybrid search query combining PGVector and TSVector",
            "AI Tool Used": "GitHub Copilot",
            "AI Output": "Provide raw SQL query combining PostgreSQL cosine similarity and ts_rank.",
            "Student’s Validation / Modification": "We evaluated the SQL query. We customized the query to apply dynamic weights (70% for vector semantic similarity and 30% for FTS term frequency). This improves search relevance when recruiters search for specific technology keywords.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week7.docx",
            "Quantitative Measure": "Build 1 optimized hybrid search query",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Threshold tuning is required to filter out unrelated low-scoring candidates.",
            "Prompt": "Write a raw SQL query for PostgreSQL combining pgvector similarity and full-text search (tsvector) to retrieve candidates.",
            "Response": "```sql\nSELECT *, (1 - (embedding <=> $1)) AS similarity, ts_rank_cd(text_search_vector, to_tsquery($2)) AS text_rank\nFROM \"CandidateCV\"\nORDER BY (similarity * 0.7 + text_rank * 0.3) DESC;\n```"
        },
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Centralize HR authorization and restrict access to interview feedback comments",
            "AI Tool Used": "Codex",
            "AI Output": "Provide authorization check code for interview result records.",
            "Student’s Validation / Modification": "We configured the FeedbackGuard. We customized it to only allow the assigned interview panel members or HR managers of the organization to modify interview feedback, restricting all other users to read-only access.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week7.docx",
            "Quantitative Measure": "Configure 1 unified authorization guard",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Recruiters can be locked out of feedback submission if interviewer IDs are mismatched in the database.",
            "Prompt": "Write guard code to ensure only assigned interviewers or HR can update candidate interview comments.",
            "Response": "In guard class, load the slot. Check if `slot.interviewerId === user.id` or `user.role === 'HR'`. Return false otherwise."
        }
    ],
    "Week8": [
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Calibrate interview feedback scores and HR candidate recommendation workflow",
            "AI Tool Used": "Codex",
            "AI Output": "Provide database schema updates to store granular scores and input validation.",
            "Student’s Validation / Modification": "We checked input constraints. We configured controllers to strictly validate that evaluation scores (technical, communication, culture fit) are integers between 1 and 5. We blocked score modifications after submission to maintain evaluation integrity.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week8.docx",
            "Quantitative Measure": "Enforce validation constraints on 3 score metrics",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Interviewers may submit duplicate reviews if the submit button is double-clicked on latency.",
            "Prompt": "Write validation code to ensure interview criteria scores are integers between 1 and 5.",
            "Response": "Validate each score parameter and throw BadRequestException if outside [1, 5] range."
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Write script to scan Gateway API routes and cross-reference with App.tsx routes",
            "AI Tool Used": "Codex",
            "AI Output": "Provide Node.js script iterating over NestJS controller metadata and printing registered routes.",
            "Student’s Validation / Modification": "We reviewed the router scanner. We executed the script to extract active backend routes and compared them with React Router declarations in App.tsx, identifying and removing 2 obsolete endpoints.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week8.docx",
            "Quantitative Measure": "Verify 39 screens against 40 API endpoints",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Dynamic path parameters can cause mismatches if string patterns are not normalized.",
            "Prompt": "Write a Node.js script analyzing NestJS controllers to list all active API routes.",
            "Response": "Use NestJS `HttpAdapterHost` to retrieve the registered router instance dynamically."
        },
        {
            "SDLC Phase": "Documentation",
            "Task / Activity": "Design screen navigation flow and system architecture diagrams",
            "AI Tool Used": "Codex",
            "AI Output": "Provide PlantUML syntax describing communication flow between Gateway and microservices.",
            "Student’s Validation / Modification": "We designed the flow diagrams. We used PlantUML to construct sequence diagrams for registration, login, and request approval flows, checking them into the repository for developer onboarding.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week8.docx",
            "Quantitative Measure": "Construct 4 detailed PlantUML diagrams",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Diagrams must be manually updated whenever microservice RPC interfaces change.",
            "Prompt": "Create a PlantUML sequence diagram for user signup and login via API Gateway and Identity Service.",
            "Response": "Define participants: User, Gateway, IdentityService, DB. Write arrows describing payload and token return."
        },
        {
            "SDLC Phase": "Documentation",
            "Task / Activity": "Build Requirement Traceability Matrix (RTM)",
            "AI Tool Used": "Codex",
            "AI Output": "Provide Markdown table template mapping business requirements to functions and tests.",
            "Student’s Validation / Modification": "We built the traceability matrix. We mapped all 22 functional requirements (FR) of the project, ensuring every specification has a corresponding API endpoint and unit test suite.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week8.docx",
            "Quantitative Measure": "Map 22 functional requirements",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Requires manual updates whenever features are modified.",
            "Prompt": "How do we set up a Requirement Traceability Matrix in Markdown to link functional requirements with APIs and tests?",
            "Response": "Create a Markdown table with columns: Req ID, Description, Database Table, API Endpoint, Service Method, Test File."
        },
        {
            "SDLC Phase": "Documentation",
            "Task / Activity": "Compile project handoff documentation and summaries",
            "AI Tool Used": "Codex, MarkItDown",
            "AI Output": "Provide structural framework for project summaries in Markdown.",
            "Student’s Validation / Modification": "We compiled the handoff documents. We replaced the placeholders with active system metrics (active endpoints, test coverage, Docker setup guide).",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week8.docx",
            "Quantitative Measure": "Compile 6 technical document guides",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Static documents risk drifting from the codebase if updates are not automated.",
            "Prompt": "Provide structure for handoff documentation detailing Gateway security, background queues, and logging configurations.",
            "Response": "Outline chapters on Gateway security, BullMQ retry queues, Pino logger configuration, and health check endpoints."
        }
    ],
    "Week9": [
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Automatically transition Campaign status to COMPLETED when target headcount is met",
            "AI Tool Used": "Gemini",
            "AI Output": "Provide logic checking hired count against campaign target to update state.",
            "Student’s Validation / Modification": "We designed the campaign completion trigger. We embedded the check within the hiring transaction. When a candidate's status updates to HIRED, we count total hired candidates. If target headcount is met, the campaign status transitions to COMPLETED within the same transaction to prevent concurrency conflicts.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week9.docx",
            "Quantitative Measure": "Configure 1 automatic status workflow",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Multiple concurrent hires might exceed targets if database row locking is not applied.",
            "Prompt": "Write a Prisma transaction and check logic to see if target headcount is met, and auto-complete the campaign.",
            "Response": "Inside hiring transaction, load campaign target and count HIRED candidates. If matched, update campaign to COMPLETED."
        },
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Design API key rotation and fallback mechanism for Gemini API",
            "AI Tool Used": "Gemini",
            "AI Output": "Provide code iterating over API key array and shifting index on rate limit error (429).",
            "Student’s Validation / Modification": "We implemented key rotation. We wrapped Gemini API calls in a retry loop. On a 429 Rate Limit error, the system cycles to the next API key. The exception is only thrown after 3 failed key rotation attempts.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week9.docx",
            "Quantitative Measure": "Automatically rotate across 3 fallback API keys",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Risk of complete quota exhaustion if all configured keys hit their daily limits.",
            "Prompt": "Implement API key rotation for Gemini in Node.js. If the current key hits a 429 rate limit, fallback and retry with a new key.",
            "Response": "```javascript\nconst keys = [process.env.GEMINI_KEY_1, process.env.GEMINI_KEY_2];\nlet keyIndex = 0;\n// In try-catch:\ncatch (err) {\n  if (err.status === 429) {\n    keyIndex = (keyIndex + 1) % keys.length;\n    // retry execution with new key index\n  }\n}\n```"
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Resolve session cookie loss on Google OAuth callback redirect",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Configure cookie SameSite and Secure attributes to allow cross-site cookie transfers.",
            "Student’s Validation / Modification": "We analyzed the cookie loss. Chrome blocks third-party cookies on redirect paths without explicit security properties. We updated session configuration at Gateway to `SameSite=Lax` and `Secure=true`, ensuring sessions are preserved on Google redirect.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week9.docx",
            "Quantitative Measure": "Resolve 1 critical authentication session bug",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Requires HTTPS in testing environment for Secure cookies to be saved.",
            "Prompt": "Why is the session cookie lost on Google OAuth redirect callback in NestJS? How do we fix it?",
            "Response": "Set `sameSite: 'lax'` or `sameSite: 'none'` (with `secure: true`) in cookie middleware options."
        },
        {
            "SDLC Phase": "Implementation",
            "Task / Activity": "Unify recruitment request filters for Department Heads and HR Managers",
            "AI Tool Used": "GitHub Copilot",
            "AI Output": "Provide React code rendering state filters and search bars.",
            "Student’s Validation / Modification": "We updated the request filter interface. We noticed display mismatches in request states between Dept Heads (creators) and HR (approvers). We refactored the filters to consume shared state enums imported from the project's contracts package.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week9.docx",
            "Quantitative Measure": "Synchronize user interfaces for 2 primary user roles",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Stale data can be displayed if page state is not re-fetched on action updates.",
            "Prompt": "Create a React table displaying recruitment requests shared by Dept Head and HR, rendering conditional buttons.",
            "Response": "Use centralized enums (DRAFT, PENDING, APPROVED) and render action buttons based on user role."
        }
    ],
    "Week10": [
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Configure Playwright E2E tests in Monorepo structure",
            "AI Tool Used": "ChatGPT",
            "AI Output": "Provide playwright.config.ts and sample test scenarios for login and CV upload.",
            "Student’s Validation / Modification": "We established E2E tests. We configured Playwright to execute in the CI/CD pipeline, and created a `globalSetup` script to clear database records before E2E runs to ensure clean test runs.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week10.docx",
            "Quantitative Measure": "Configure 3 E2E test suites",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "E2E tests can be flaky if UI render speeds drop below Playwright's timeout limits.",
            "Prompt": "Write a Playwright config file for Turborepo monorepo, specifying test directories and server command.",
            "Response": "```javascript\nimport { defineConfig } from '@playwright/test';\nexport default defineConfig({\n  testDir: './e2e',\n  use: { baseURL: 'http://localhost:5173' },\n  webServer: { command: 'npm run dev', url: 'http://localhost:5173' },\n});\n```"
        },
        {
            "SDLC Phase": "Testing",
            "Task / Activity": "Build API integration test runner using Postman collections and Newman",
            "AI Tool Used": "GitHub Copilot",
            "AI Output": "Provide Node.js script executing Postman JSON collections via Newman.",
            "Student’s Validation / Modification": "We integrated API tests. We customized the runner to output HTML reports and exit with non-zero codes on test failures to block broken CI builds.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week10.docx",
            "Quantitative Measure": "Configure 1 automated API test runner",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "Static parameters in Postman collections can become stale as database schema evolves.",
            "Prompt": "Write a Node.js script using Newman to run a Postman collection and export results.",
            "Response": "```javascript\nconst newman = require('newman');\nnewman.run({\n  collection: require('./rms.postman_collection.json'),\n  reporters: 'cli',\n}, function (err) { if (err) { throw err; } });\n```"
        },
        {
            "SDLC Phase": "Automation",
            "Task / Activity": "Automate task assignee updates using openpyxl and GitHub API",
            "AI Tool Used": "Gemini",
            "AI Output": "Provide Python script reading Excel, comparing values, and calling GitHub API using urllib.request.",
            "Student’s Validation / Modification": "We tested the sync script. We verified GitHub account mapping and added pagination and rate limit controls to handle batch updates of 50+ issues without being throttled.",
            "Evidence / Link": "https://github.com/vvtPhongdev/SE20A05Group7RMS/blob/docs/project-report-and-workflow-updates/docs/Report/AI_Evidence/Week10.docx",
            "Quantitative Measure": "Sync assignee state for 58 issues",
            "Value Added (1-5)": 5,
            "Risks / Limitations Observed": "GitHub access token can expire or lack write permissions to the repository.",
            "Prompt": "Write a Python script loading Excel with openpyxl, extracting issue IDs, matching assignees, and updating GitHub via PATCH.",
            "Response": "Define `update_issues.py` utilizing load_workbook, urllib.request, and JSON payload."
        }
    ]
}

def create_markdown_evidence():
    print("Generating detailed AI Evidence markdown files in English...")
    os.makedirs("docs/Report/AI_Evidence", exist_ok=True)
    for week_name, tasks in WEEKLY_DATA.items():
        filename = f"docs/Report/AI_Evidence/{week_name}.md"
        with open(filename, "w", encoding="utf-8") as f:
            for idx, task in enumerate(tasks):
                task_id = idx + 1
                f.write(f"---\n\n")
                f.write(f"## Task {task_id}: {task['Task / Activity']}\n\n")
                f.write(f"### 1. Prompt Sent to AI (Prompt)\n\n")
                f.write(f"```text\n{task['Prompt']}\n```\n\n")
                f.write(f"### 2. Detailed AI Response\n\n")
                f.write(f"{task['Response']}\n\n")
                f.write(f"### 3. Student's Validation, Selection & Source Code Customization\n\n")
                f.write(f"**Verification Description:**\n")
                f.write(f"{task['Student’s Validation / Modification']}\n\n")
                f.write(f"**Audit Information:**\n")
                f.write(f"- **SDLC Phase:** {task['SDLC Phase']}\n")
                f.write(f"- **AI Tool Used:** {task['AI Tool Used']}\n")
                f.write(f"- **Quantitative Measure:** {task['Quantitative Measure']}\n")
                f.write(f"- **Risks & Limitations Observed:** {task['Risks / Limitations Observed']}\n\n")
        print(f"  Created detailed {filename}")

def update_excel_report():
    print("Updating SWP_SE20A05_Group7_ReportAI.xlsx in English...")
    wb = openpyxl.load_workbook("docs/Report/SWP_SE20A05_Group7_ReportAI.xlsx")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    font_data = Font(name='Segoe UI', size=11, bold=False)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center')
    
    for sheet_name, tasks in WEEKLY_DATA.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: sheet {sheet_name} not found, creating it...")
            wb.create_sheet(sheet_name)
            
        sheet = wb[sheet_name]
        
        # Unmerge all merged ranges in this sheet to prevent MergedCell read-only exceptions
        merged_ranges = list(sheet.merged_cells.ranges)
        for rng in merged_ranges:
            sheet.unmerge_cells(str(rng))
            
        # Clear rows from Row 2 to 100, columns 1 to 20 to completely erase any previous data
        for r in range(2, 101):
            for c in range(1, 21):
                cell = sheet.cell(r, c)
                cell.value = None
                cell.hyperlink = None
                    
        # Write tasks
        for idx, task in enumerate(tasks):
            r = idx + 2
            sheet.cell(r, 1).value = idx + 1
            sheet.cell(r, 2).value = task["SDLC Phase"]
            sheet.cell(r, 3).value = task["Task / Activity"]
            sheet.cell(r, 4).value = task["AI Tool Used"]
            sheet.cell(r, 5).value = task["AI Output"]
            sheet.cell(r, 6).value = task["Student’s Validation / Modification"]
            sheet.cell(r, 7).value = task["Evidence / Link"]
            sheet.cell(r, 8).value = task["Quantitative Measure"]
            sheet.cell(r, 9).value = task["Value Added (1-5)"]
            sheet.cell(r, 10).value = task["Risks / Limitations Observed"]
            
            # Formatting
            for c in range(1, 11):
                cell = sheet.cell(r, c)
                cell.font = font_data
                cell.border = thin_border
                if c in [1, 4, 9]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
        # Adjust row heights and column widths
        sheet.row_dimensions[1].height = 25
        for idx in range(len(tasks)):
            sheet.row_dimensions[idx + 2].height = 65
            
        for col in sheet.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter in ['A', 'D', 'I']:
                sheet.column_dimensions[col_letter].width = 12
            elif col_letter in ['B', 'C', 'E', 'F', 'G', 'H', 'J']:
                sheet.column_dimensions[col_letter].width = 30
                
    wb.save("docs/Report/SWP_SE20A05_Group7_ReportAI.xlsx")
    print("Excel save complete.")

if __name__ == "__main__":
    create_markdown_evidence()
    update_excel_report()
